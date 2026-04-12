"""Unit tests for the non-GUI core modules."""

from __future__ import annotations

from pathlib import Path

import numpy as np
import pytest

from app.core import edge_detector, excel_exporter
from app.core.measurement import Layer, Measurement
from app.core.scale_calibrator import ScaleCalibrator


# ----------------------------------------------------------------------
# ScaleCalibrator
# ----------------------------------------------------------------------


def test_scale_calibrator_basic():
    cal = ScaleCalibrator()
    assert not cal.is_calibrated
    ratio = cal.set_reference((0.0, 0.0), (100.0, 0.0), real_mm=10.0)
    assert cal.is_calibrated
    assert ratio == pytest.approx(0.1)
    assert cal.to_mm(50) == pytest.approx(5.0)


def test_scale_calibrator_rejects_zero_length():
    cal = ScaleCalibrator()
    with pytest.raises(ValueError):
        cal.set_reference((5.0, 5.0), (5.0, 5.0), real_mm=10.0)


def test_scale_calibrator_rejects_non_positive_mm():
    cal = ScaleCalibrator()
    with pytest.raises(ValueError):
        cal.set_reference((0.0, 0.0), (10.0, 0.0), real_mm=0.0)


# ----------------------------------------------------------------------
# edge_detector
# ----------------------------------------------------------------------


def _make_synthetic_section(layer_heights, width=40, pad=6):
    """Create a BGR image with horizontal stripes of given pixel heights.

    The image is padded on top and bottom with a uniform background so that
    the outermost layers also have a detectable edge on both sides.
    """
    total_h = sum(layer_heights) + 2 * pad
    img = np.full((total_h, width, 3), 255, dtype=np.uint8)
    y = pad
    greys = [30, 200, 90, 230, 50, 170, 110]
    for i, h in enumerate(layer_heights):
        img[y : y + h, :, :] = greys[i % len(greys)]
        y += h
    return img


def test_detect_layers_recovers_known_thicknesses():
    heights = [20, 35, 25]
    img = _make_synthetic_section(heights)
    h, w, _ = img.shape
    x = w // 2
    layers = edge_detector.detect_layers(img, (x, 0), (x, h - 1))
    assert len(layers) == len(heights)
    for layer, expected in zip(layers, heights):
        assert layer.thickness_px == pytest.approx(expected, abs=2)


def test_detect_layers_returns_empty_for_uniform_image():
    img = np.full((80, 40, 3), 128, dtype=np.uint8)
    layers = edge_detector.detect_layers(img, (20, 0), (20, 79))
    assert layers == []


def test_layers_from_boundaries_rebuilds_layers():
    layers = edge_detector.layers_from_boundaries([10, 30, 60, 61])
    # The 60→61 boundary is below min_thickness_px (default 2) so it is dropped.
    assert [round(layer.thickness_px) for layer in layers] == [20, 30]
    assert layers[0].name == "Layer 1"
    assert layers[1].name == "Layer 2"


# ----------------------------------------------------------------------
# Measurement + excel export
# ----------------------------------------------------------------------


def test_measurement_recompute_applies_mm_ratio():
    m = Measurement(mm_per_pixel=0.05)
    m.replace_layers(
        [
            Layer("Layer 1", 0, 20, 20, 0),
            Layer("Layer 2", 20, 50, 30, 0),
        ]
    )
    assert m.layers[0].thickness_mm == pytest.approx(1.0)
    assert m.layers[1].thickness_mm == pytest.approx(1.5)


def test_excel_export_roundtrip(tmp_path: Path):
    m = Measurement(
        image_path=tmp_path / "sample.png",
        mm_per_pixel=0.1,
    )
    m.replace_layers(
        [
            Layer("Top", 0, 10, 10, 0),
            Layer("Mid", 10, 40, 30, 0),
            Layer("Bot", 40, 60, 20, 0),
        ]
    )
    out = excel_exporter.export(tmp_path / "out.xlsx", m)
    assert out.exists()

    from openpyxl import load_workbook

    wb = load_workbook(out)
    ws = wb.active
    # Metadata rows
    assert ws.cell(row=1, column=1).value == "Image"
    assert ws.cell(row=2, column=1).value == "mm per pixel"
    assert float(ws.cell(row=2, column=2).value) == pytest.approx(0.1)
    # Header row at 4, body starts at 5.
    headers = [ws.cell(row=4, column=c).value for c in range(1, 6)]
    assert headers == [
        "Layer",
        "Top (px)",
        "Bottom (px)",
        "Thickness (px)",
        "Thickness (mm)",
    ]
    names = [ws.cell(row=r, column=1).value for r in range(5, 8)]
    assert names == ["Top", "Mid", "Bot"]
    mm_values = [ws.cell(row=r, column=5).value for r in range(5, 8)]
    assert mm_values == pytest.approx([1.0, 3.0, 2.0])
