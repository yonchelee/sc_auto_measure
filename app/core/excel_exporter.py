"""Export a :class:`Measurement` to an .xlsx workbook."""

from __future__ import annotations

from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .measurement import Measurement


COLUMNS = [
    "Layer",
    "Top (px)",
    "Bottom (px)",
    "Thickness (px)",
    "Thickness (mm)",
]


def measurement_to_dataframe(measurement: Measurement) -> pd.DataFrame:
    rows = []
    for layer in measurement.layers:
        rows.append(
            {
                "Layer": layer.name,
                "Top (px)": round(layer.y_top_px, 2),
                "Bottom (px)": round(layer.y_bottom_px, 2),
                "Thickness (px)": round(layer.thickness_px, 2),
                "Thickness (mm)": round(layer.thickness_mm, 4),
            }
        )
    return pd.DataFrame(rows, columns=COLUMNS)


def _safe_sheet_name(raw: str) -> str:
    bad = set('[]:*?/\\')
    cleaned = "".join("_" if ch in bad else ch for ch in raw)
    return (cleaned or "Measurement")[:31]


def export(
    path: Path,
    measurement: Measurement,
    image_path: Optional[Path] = None,
) -> Path:
    """Write ``measurement`` to ``path`` and return the resolved path."""
    out_path = Path(path)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    df = measurement_to_dataframe(measurement)
    sheet_name = _safe_sheet_name(
        (image_path or measurement.image_path or Path("Measurement")).stem
    )

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Leave top rows for metadata, write table starting at row 4.
        df.to_excel(writer, sheet_name=sheet_name, startrow=3, index=False)

    # Post-process with openpyxl: add metadata and basic formatting.
    wb = load_workbook(out_path)
    ws = wb[sheet_name]

    meta_font = Font(bold=True)
    ws.cell(row=1, column=1, value="Image").font = meta_font
    ws.cell(row=1, column=2, value=str(image_path or measurement.image_path or ""))
    ws.cell(row=2, column=1, value="mm per pixel").font = meta_font
    ws.cell(
        row=2,
        column=2,
        value=float(measurement.mm_per_pixel) if measurement.mm_per_pixel else "n/a",
    )

    for col_idx, name in enumerate(COLUMNS, start=1):
        header_cell = ws.cell(row=4, column=col_idx)
        header_cell.font = meta_font
        width = max(len(name), 12) + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    wb.save(out_path)
    return out_path
